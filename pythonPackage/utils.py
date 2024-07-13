import pandas as pd
import openpyxl
import gspread, requests, os
import yfinance as yf
import json
from datetime import datetime, timedelta
from pythonPackage.constants import *
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request


# Common Utility Functions
def update_env_file(key, value, env_file):
    print("updating env file")
    with open(env_file, 'r') as file:
        lines = file.readlines()    

    updated_lines = []
    for line in lines:
        if line.strip().startswith(f"{key}="):
            line = f"{key}={value}\n"
        updated_lines.append(line)

    print(updated_lines)
    with open(env_file, 'w') as file:
        file.writelines(updated_lines)

def get_args_and_input(args, excel_file_name, spreadsheet_id, env_file):
    
    input_file = None
    value = ''
    key = 'EXCEL_FILE_NAME'
    credentials=None
    print("Args length: ", len(args))
    print("Args: ", args)
    if len(args) > 4:
        credentials = json.loads(args[4])       
    if len(args) > 3:
        value = args[3]
    if len(args) > 2:
        typ = args[2].lower()
        if typ == 'sheets':
            key = 'SPREADSHEET_ID'
        input_file = args[1]
        if input_file == 'None':
            input_file = None
    else:
        if len(args) > 1:
            input_file = args[1]
        else:
            input_file = input("Please enter the absolute path of the file downloaded from Zerodha: ")
        print("\nPlease select 'excel' or 'sheets, leave empty to use excel as default")
        typ = input("Enter your choice: ")
        typ = typ.lower()
        if (typ == 'excel' or typ == ''):
            print(
                f"{excel_file_name} is the default file, enter name below if you wish to change it, leave empty otherwise")
            value = input("Enter your choice: ")

        elif (typ== 'sheets'):
            print(f"{spreadsheet_id} is the default google sheet, enter spreadsheet_id below if you wish to change it, leave empty otherwise")
            value = input("Enter your choice: ")
            key = 'SPREADSHEET_ID'

    print("going to update env file", value)
    if value is not None and value != '':
        update_env_file(key, value, env_file)

    return input_file, typ, credentials

def get_sheets_and_data(typ, credentials_file, spreadsheet_id, spreadsheet_file, credentials=None):
    if typ == 'sheets':
        spreadsheet = authenticate_and_get_sheets(
            credentials_file, spreadsheet_id, credentials)
        worksheets = spreadsheet.worksheets()
        sheet_names = [worksheet.title for worksheet in worksheets]
        raw_data = read_data_from_sheets(spreadsheet, sheet_names[0])
    else:
        spreadsheet = openpyxl.load_workbook(spreadsheet_file)
        sheet_names = spreadsheet.sheetnames
        raw_data = read_data_from_excel(spreadsheet, sheet_names[0])

    return spreadsheet, sheet_names, raw_data

def data_already_exists(raw_data, input_data):
    input_data.reset_index(drop=True, inplace=True)
    if not raw_data.empty:
        is_duplicate = raw_data[
            (raw_data[Raw_constants.DATE] == input_data[Raw_constants.DATE][0]) &
            (raw_data[Raw_constants.NAME] == input_data[Raw_constants.NAME][0])
        ].shape[0] > 0

        if is_duplicate:
            print("Orders Data Already Exists in the file, Exiting...")
            exit()

# Functions Specific To Sheets

def authenticate_and_get_sheets(credentials_file, spreadsheet_id, credentials=None):
    print("Yet to enter")
    if credentials is not None:
        print("Entering here finally")
        try:
            credentials_obj = Credentials.from_authorized_user_info(credentials)
            if credentials_obj and credentials_obj.expired and credentials_obj.refresh_token:
                credentials_obj.refresh(Request())
            else:
                raise ValueError("OAuth Credentials are not available or invalid")
            
            gc = gspread.authorize(credentials_obj)
            spreadsheet = gc.open_by_key(spreadsheet_id)
            return spreadsheet
        except Exception as e:
            print(f"Unable to authorize spreadsheet {e}, exiting...")
            exit()
            
    else:
        gc = gspread.service_account(filename=credentials_file)
        spreadsheet = gc.open_by_key(spreadsheet_id)
        return spreadsheet
    
        
def read_data_from_sheets(spreadsheet, sheet_name):
    sheet = spreadsheet.worksheet(sheet_name)
    data = sheet.get_all_values()
    if len(data) == 0:
        return pd.DataFrame()
    df = pd.DataFrame(data[1:], columns=data[0])
    return df

def format_background_sheets(spreadsheet ,sheet, cell_range):
    requests = [{
        "updateCells": {
            "range": {
                "sheetId": sheet.id,
                "startRowIndex": int(cell_range.split(':')[0][1:]) - 1,
                "endRowIndex": int(cell_range.split(':')[1][1:]),
                "startColumnIndex": ord(cell_range.split(':')[0][0].upper()) - 65,
                "endColumnIndex": ord(cell_range.split(':')[1][0].upper()) - 64,
            },
            "fields": "userEnteredFormat.backgroundColor",
        }
    }]
    spreadsheet.batch_update({"requests":requests})

def display_and_format_sheets(sheet, data):
    data = data.round(4)
    headers = data.columns.tolist() 
    data = data.values.tolist() 

    sheet.update('A1', [headers])
    sheet.insert_rows(data, 2)
    
    # Add formatting to headers
    num_columns = len(headers)
    header_range = f'A1:{chr(64 + num_columns)}1'
    sheet.format(header_range, {"textFormat": {"bold": True}})

def initialize_sheets(spreadsheet, sheet_name):
    sheet = spreadsheet.worksheet(sheet_name)
    sheet.clear() 
    format_background_sheets(spreadsheet, sheet, CELL_RANGE)    
    return sheet

def transDetails_formatting_sheets(spreadsheet, sheet):
    worksheet_data = sheet.get_all_values()
    requests = []
    for row_number, row_data in  enumerate(worksheet_data[1:], start=2):
        transaction_type = row_data[6]
        if transaction_type == BUY:
            background_color = (0.8, 0.9, 1)
        else:
            background_color = (1, 0.8, 0.8)
        format_request = get_backgroundColor_formatting_request(sheet, row_number, row_data, background_color)
        requests.append(format_request)
    if len(requests) > 0:
        spreadsheet.batch_update({"requests": requests})

def shareProfitLoss_formatting_sheets(spreadsheet, sheet):
    worksheet_data = sheet.get_all_values()
    requests = []
    for row_number, row_data in  enumerate(worksheet_data[1:], start=2):
        remaining_shares = int(row_data[7])
        profit = float(row_data[9])
        if remaining_shares == 0:
            if profit > 0:
                background_color = (0.8, 0.9, 1)
            else:
                background_color = (1, 0.8, 0.8)
            format_request = get_backgroundColor_formatting_request(sheet,row_number, row_data, background_color)
            requests.append(format_request)
    if len(requests) > 0:
        spreadsheet.batch_update({"requests": requests})

def dailyProfitLoss_formatting_sheets(spreadsheet, sheet):
    worksheet_data = sheet.get_all_values()
    requests = []
    for row_number, row_data in  enumerate(worksheet_data[1:], start=2):
        if row_data[10] != "":
            if float(row_data[10]) > 0.0:
                background_color = (0.8, 0.9, 1)
            else:
                background_color = (1, 0.8, 0.8)
            format_request = get_backgroundColor_formatting_request(sheet,row_number, row_data, background_color)
            requests.append(format_request)
    if len(requests) > 0:
        spreadsheet.batch_update({"requests": requests})

def taxation_formatting_sheets(spreadsheet, sheet):
    worksheet_data = sheet.get_all_values()
    requests = []
    for row_number, row_data in  enumerate(worksheet_data[1:], start=2):
        if (float(row_data[2]) + float(row_data[3]) + float(row_data[4])) >= 0:
            background_color = (0.8, 0.9, 1)
        else:
            background_color = (1, 0.8, 0.8)
        format_request = get_backgroundColor_formatting_request(sheet,row_number, row_data, background_color)
        requests.append(format_request)
    if len(requests) > 0:
        spreadsheet.batch_update({"requests": requests})



# Functions Specific To Excel

def read_data_from_excel(spreadsheet_file, sheet_name):
    df = pd.read_excel(spreadsheet_file, sheet_name=sheet_name)
    return df

def format_background_excel(sheet, cell_range):
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(fill_type=None)

def display_and_format_excel(sheet, data):
    data = data.round(4)
    try:
        data[Raw_constants.DATE] = data[Raw_constants.DATE].dt.strftime(DATE_FORMAT)
    except Exception as e:
        pass
    headers = data.columns.tolist() 
    data = data.values.tolist() 
    sheet.append(headers)
    for row in data:
        sheet.append(row)
    for cell in sheet['1']:
        cell.font = openpyxl.styles.Font(bold=True)

def initialize_excel(spreadsheet, sheet_name):
    sheet = spreadsheet[sheet_name]
    sheet.clear()
    format_background_excel(sheet, CELL_RANGE)
    return sheet

def transDetails_formatting_excel(sheet):
    cell_range = CELL_RANGE
    redFill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    blueFill = openpyxl.styles.PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
    for row in sheet[cell_range]:
        if row[0].value is None or row[0].value == '':
            break
        if row[4].value == BUY:
            for cell in row:
                cell.fill = blueFill
        elif row[4].value == SELL:
            for cell in row:
                cell.fill = redFill

def shareProfitLoss_formatting_excel(sheet):
    cell_range = CELL_RANGE
    redFill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    blueFill = openpyxl.styles.PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
    for row in sheet[cell_range]:
        if row[0].value is None or row[0].value == '':
            break
        remaining_shares = int(row[7].value)
        profit = float(row[9].value)
        if remaining_shares == 0:
            if profit >= 0.0:
                for cell in row:
                    cell.fill = blueFill
            else:
                for cell in row:
                    cell.fill = redFill

def dailyProfitLoss_formatting_excel(sheet):
    cell_range = CELL_RANGE
    redFill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    blueFill = openpyxl.styles.PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
    for row in sheet[cell_range]:
        if row[0].value is None or row[0].value == '':
            break
        if row[9].value != '':
            if float(row[9].value) > 0.0:
                for cell in row:
                    cell.fill = blueFill
            else :
                for cell in row:
                    cell.fill = redFill
                    
def taxation_formatting_excel(sheet):
    cell_range = CELL_RANGE
    redFill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    blueFill = openpyxl.styles.PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
    for row in sheet[cell_range]:
        if row[0].value is None or row[0].value == '':
            break
        if row[9].value != '':
            if float(row[2].value) + float(row[3].value) + float(row[4].value) > 0.0:
                for cell in row:
                    cell.fill = blueFill
            elif float(row[2].value) + float(row[3].value) + float(row[4].value) < 0.0:
                for cell in row:
                    cell.fill = redFill
    pass


# Functions Specific To Data

def initialize_data(data, extraCols=[], sortList=[]):
    if len(extraCols) > 0:
        data = data.drop(extraCols,axis=1)
    
    for header in data.columns.tolist():
        if is_formatable(data[header].iloc[0]):
            data[header] = pd.to_numeric(data[header])
    data[Raw_constants.DATE] = pd.to_datetime(data[Raw_constants.DATE], format=DATE_FORMAT)
    
    if len(sortList) > 0:
        data.sort_values(by=sortList, inplace=True)
    data = data.round(4)
    try:
        data[Raw_constants.DATE] = data[Raw_constants.DATE].dt.strftime(DATE_FORMAT)
    except Exception as e:
        pass
    return data

"""
Input Format:
    input_data: "Time","Type","Instrument","Product","Qty.","Avg. price","Status"

Output Format:
    df: Date, Name, Price, Quantity, Net Amount, Stock Exchange
"""
def format_input_data(input_data):
    input_data.drop(
        input_data[input_data[Order_constants.STATUS] != COMPLETE].index, inplace=True)
    input_data[Order_constants.QUANTITY] = input_data.apply(
        get_order_quantity, axis=1)

    constants_dict = {key: value for key, value in vars(Raw_constants).items(
    ) if (isinstance(value, str) and not value.startswith('python'))}
    df = pd.DataFrame(columns=list(constants_dict.values()))

    df[Raw_constants.DATE] = input_data[Order_constants.DATE].apply(
        lambda x: get_order_date(x))
    df[Raw_constants.NAME] = input_data[Order_constants.NAME]
    df[Raw_constants.PRICE] = input_data[Order_constants.PRICE]
    df[Raw_constants.QUANTITY] = input_data[Order_constants.QUANTITY]
    df[Raw_constants.NET_AMOUNT] = df.apply(get_net_amount, axis=1)
    df[Raw_constants.STOCK_EXCHANGE] = BSE
    return df

"""
Input Format: 
    input_data: symbol, isin, trade_date, exchange, segment, series, trade_type, auction, quantity, price, trade_id, order_id, order_execution_time

Output Format:
    df: Date, Name, Price, Quantity, Net Amount, Stock Exchange 
"""
def format_add_data(input_data):
    input_data[Data_constants.QUANTITY] = input_data.apply(
        get_data_quantity, axis=1)
    
    constants_dict = {key: value for key, value in vars(Raw_constants).items(
    ) if (isinstance(value, str) and not value.startswith('python'))}
    df = pd.DataFrame(columns=list(constants_dict.values()))
    
    df[Raw_constants.DATE] = input_data[Data_constants.DATE].apply(
        lambda x: get_data_date(x))
    df[Raw_constants.NAME] = input_data[Data_constants.NAME]
    df[Raw_constants.PRICE] = input_data[Data_constants.PRICE]
    df[Raw_constants.QUANTITY] = input_data[Data_constants.QUANTITY]
    df[Raw_constants.NET_AMOUNT] = df.apply(get_net_amount, axis=1)
    df[Raw_constants.STOCK_EXCHANGE] = input_data[Data_constants.STOCK_EXCHANGE]
    return df

# Utility Functions

def get_valid_path(path):
    if path is None:
        return None
    if not os.path.exists(path):
        print("The Provided Path for file downloaded from Zerodha does not exist!")
        path = input("Enter correct Absolute Path, including /home: ")
        path = get_valid_path(path)   
    return path

def get_order_quantity(row):
    quantity = row[Order_constants.QUANTITY]
    val = quantity.split('/')[0]
    if row[Order_constants.TYPE] == SELL:
        val = '-' + val
    return val

def get_data_quantity(row):
    quantity = row[Data_constants.QUANTITY]
    val = str(quantity).split('.')[0]
    if str(row[Data_constants.TYPE]).upper() == SELL:
        val = '-' + val
    return val

def get_order_date(date):
    date_obj = datetime.strptime(date, ORDER_TIME_FORMAT)
    date_str = datetime.strftime(date_obj, DATE_FORMAT)
    return date_str

def get_data_date(date):
    date_obj = datetime.strptime(date, DATA_TIME_FORMAT)
    date_str = datetime.strftime(date_obj, DATE_FORMAT)
    return date_str

    

def get_net_amount(row):
    val = int(row[Raw_constants.QUANTITY]) * float(row[Raw_constants.PRICE])
    return str(val)

def update_transaction_type(x):
    if x[0] != '-':
        return BUY
    return SELL

def update_intraday_count(data):
    infoMap = {}
    rowData = {}
    sumData = {}
    grouped_data = data.groupby(
        [Raw_constants.NAME, TransDetails_constants.TRANSACTION_TYPE, Raw_constants.DATE])
    for (name, transaction_type, date), group in grouped_data:
        if name not in infoMap:
            infoMap[name] = {}
            rowData[name] = {}
            sumData[name] = {}
        if date not in infoMap[name]:
            infoMap[name][date] = {}
            rowData[name][date] = []
            sumData[name][date] = 0
        if transaction_type not in infoMap[name][date]:
            infoMap[name][date][transaction_type] = group.values.tolist()

        if transaction_type == SELL:
            buyCnt = 0

            if BUY in infoMap[name][date]:
                for x in infoMap[name][date][BUY]:
                    buyCnt += abs(x[3])
            for x in infoMap[name][date][transaction_type]:
                sellCnt = min(abs(x[3]), buyCnt)
                buyCnt -= sellCnt
                rowData[name][date].append(sellCnt)
                sumData[name][date] += sellCnt

    data[TransDetails_constants.INTRADAY_COUNT] = data.apply(
        calculate_intra_sell_count, axis=1, args=(rowData,))
    data[TransDetails_constants.INTRADAY_COUNT] = data.apply(
        calculate_intra_buy_count, axis=1, args=(sumData,))

def calculate_intra_sell_count(row, rowData):
    name = row[Raw_constants.NAME]
    date = row[Raw_constants.DATE]
    transaction_type = row[TransDetails_constants.TRANSACTION_TYPE]

    if transaction_type == BUY:
        return 0

    if name not in rowData or date not in rowData[name]:
        return 0

    if len(rowData[name][date]) > 0:
        intraCount = rowData[name][date][0]
        rowData[name][date].pop(0)
        return intraCount

def calculate_intra_buy_count(row, sumData):
    name = row[Raw_constants.NAME]
    date = row[Raw_constants.DATE]
    transaction_type = row[TransDetails_constants.TRANSACTION_TYPE]

    if name not in sumData or date not in sumData[name]:
        return 0

    if transaction_type == SELL:
        return row[TransDetails_constants.INTRADAY_COUNT]

    intraCount = min(sumData[name][date], row[Raw_constants.QUANTITY])
    sumData[name][date] -= intraCount
    return intraCount

def calculate_stt(row):
    intraDay_charges, delivery_charges = 0, 0
    delivery_charges = (abs(row[Raw_constants.QUANTITY]) -
                        row[TransDetails_constants.INTRADAY_COUNT]) * 0.001 * row[Raw_constants.PRICE]
    if row[TransDetails_constants.TRANSACTION_TYPE] == SELL:
        intraDay_charges = row[TransDetails_constants.INTRADAY_COUNT] * \
            0.00025 * row[Raw_constants.PRICE]
    return intraDay_charges + delivery_charges

def calculate_transaction_charges(row):
    if row[Raw_constants.STOCK_EXCHANGE] == BSE:
        return abs(row[Raw_constants.NET_AMOUNT]) * 0.0000375
    return abs(row[Raw_constants.NET_AMOUNT]) * 0.0000335

def calculate_stamp_duty(row):
    if row[TransDetails_constants.TRANSACTION_TYPE] == SELL:
        return 0

    intraDay_charges, delivery_charges = 0, 0
    delivery_charges = (abs(row[Raw_constants.QUANTITY]) -
                        row[TransDetails_constants.INTRADAY_COUNT]) * 0.00015 * row[Raw_constants.PRICE]
    intraDay_charges = row[TransDetails_constants.INTRADAY_COUNT] * \
        0.00003 * row[Raw_constants.PRICE]
    return intraDay_charges + delivery_charges

def calculate_dp_charges(row, dp_data):
    name = row[Raw_constants.NAME]
    date = row[Raw_constants.DATE]
    if row[TransDetails_constants.TRANSACTION_TYPE] == SELL and abs(row[Raw_constants.QUANTITY]) > row[TransDetails_constants.INTRADAY_COUNT]:
        if name not in dp_data:
            dp_data[name] = {}
        if date not in dp_data[name]:
            dp_data[name][date] = True
            return 15.93
    return 0

def calculate_brokerage(row):
    if row[TransDetails_constants.INTRADAY_COUNT] > 0:
        return min(((row[TransDetails_constants.NET_AMOUNT] * row[TransDetails_constants.INTRADAY_COUNT])/row[TransDetails_constants.QUANTITY]) * 0.0003, 20)
    # val = min(data[TransDetails_constants.NET_AMOUNT] * 0.0003, 20)
    return 0

def compare_dates(d1, d2):
    dt1 = datetime.strptime(d1, DATE_FORMAT)
    dt2 = datetime.strptime(d2, DATE_FORMAT)

    if dt1 == dt2:
        return 0
    elif dt1 > dt2:
        return 1

    return -1

def calculate_average_cost_of_sold_shares(infoMap):
    sold_list = infoMap[SELL]
    buy_list = infoMap[BUY]

    j = 0
    price = 0
    intraCount = 0
    delCount = 0
    counter = 0

    # Calculating for IntraDay Orders
    for i in range(0, len(sold_list)):
        sold_list[i][3] = abs(sold_list[i][3])
        sold_list[i][6] = abs(sold_list[i][6])
        if j >= len(buy_list):
            break
        if compare_dates(buy_list[j][0], sold_list[i][0]) == 0:
            if buy_list[j][3] < sold_list[i][3]:
                intraCount += buy_list[j][3]
                sold_list[i][3] -= buy_list[j][3]
                price += buy_list[j][6]
                buy_list[j][3] = 0
                buy_list[j][6] = 0
                j += 1
                i -= 1
            elif buy_list[j][3] > sold_list[i][3]:
                price += ((buy_list[j][6] * sold_list[i][3])/buy_list[j][3])
                buy_list[j][6] -= ((buy_list[j][6] *
                                    sold_list[i][3])/buy_list[j][3])
                buy_list[j][3] -= sold_list[i][3]
                intraCount += sold_list[i][3]
                sold_list[i][3] = 0
            else:
                intraCount += sold_list[i][3]
                sold_list[i][3] = 0
                price += buy_list[j][6]
                buy_list[j][3] = 0
                buy_list[j][6] = 0
                j += 1
        elif compare_dates(buy_list[j][0], sold_list[i][0]) < 0:
            i -= 1
            j += 1

    # Calculating for Delivery Orders
    for x in sold_list:
        delCount += x[3]
    for x in buy_list:
        if x[3] <= (delCount - counter):
            price += x[6]
            counter += x[3]
        else:
            price += (x[6] * (delCount - counter))/x[3]
            counter = delCount
    counter += intraCount

    return price/counter

def is_float(input_value):
    try:
        float(input_value)
        return True
    except ValueError:
        return False

def is_formatable(input_value):
    if isinstance(input_value,str) and input_value.isnumeric():
        return True
    return is_float(input_value)

def get_backgroundColor_formatting_request(sheet, row_number, row_data, background_color):
    format_request = {
        "repeatCell": {
            "range": {
                "sheetId": sheet.id,
                "startRowIndex": row_number - 1,
                "endRowIndex": row_number,
                "startColumnIndex": 0,
                "endColumnIndex": len(row_data)
            },
            "cell": {
                "userEnteredFormat": {
                    "backgroundColor": {
                        "red": background_color[0],
                        "green": background_color[1],
                        "blue": background_color[2]
                    }
                }
            },
            "fields": "userEnteredFormat.backgroundColor"
        }
    }
    return format_request

def get_spl_row():
    context = {
        ShareProfitLoss_constants.AVERAGE_SALE_PRICE: 0,
        ShareProfitLoss_constants.AVERAGE_BUY_PRICE: 0,
        ShareProfitLoss_constants.AVERAGE_COST_OF_SOLD_SHARES: 0,
        ShareProfitLoss_constants.SHARES_SOLD: 0,
        ShareProfitLoss_constants.SHARES_BOUGHT: 0,
        ShareProfitLoss_constants.DATE: DEFAULT_DATE,
        ShareProfitLoss_constants.CURRENT_INVESTMENT: 0,
        ShareProfitLoss_constants.TOTAL_INVESTMENT: 0,
    }
    return context

def get_taxation_row():
    context = {
        Taxation_constants.DATE: DEFAULT_DATE,
        Taxation_constants.LTCG: 0.0,
        Taxation_constants.STCG: 0.0,
        Taxation_constants.INTRADAY_INCOME: 0.0,
    }
    return context

def get_prizing_details_yfinance(date, name):
    name = name.upper()
    nse_name = name + DOT_NS
    bse_name = name + DOT_BO
    output = [0,0,0,0,0]
    try:
        data = yf.download(nse_name, start=date,end=min(datetime.now(), date+timedelta(days=1)), period='10y')
        output = [data['Open'].iloc[0], data['High'].iloc[0], data['Low'].iloc[0], data['Close'].iloc[0], data['Volume'].iloc[0]]
        return [float(x) for x in output]
    except Exception as e:
        try:
            data = yf.download(bse_name, start=date,end=min(datetime.now(), date+timedelta(days=1)), period='5d')
            output = [data['Open'].iloc[0], data['High'].iloc[0], data['Low'].iloc[0], data['Close'].iloc[0], data['Volume'].iloc[0]]
            return [float(x) for x in output]
        except Exception as e:
            print("Encountered excpetion using yfinance API: ", e)
    return output

def get_prizing_details_alphaVantage(stock_name, function):
    STOCK_API_KEY = os.getenv('STOCK_API_KEY')
    url = f'https://www.alphavantage.co/query?function={function}&symbol={stock_name}&apikey={STOCK_API_KEY}&outputsize=full'
    output = []
    try:
        response = requests.get(url)
        data = response.json()
        if function == 'TIME_SERIES_DAILY':
            time_data = data['Time Series (Daily)']
            return time_data
        else:
            data = data['Global Quote']
            output = [data['02. open'], data['03. high'], data['04. low'], data['05. price'], data['06. volume']]
            return [float(x) for x in output]
    except Exception as e:
        print(f"Unable to get Share pricing details using Alpha-Vintage, Fallback to Yfinance for stock {stock_name} API: {e}")
        data = get_prizing_details_yfinance(datetime.now() - timedelta(days=1), stock_name)
        return data



# Functions for Google sheets & Excel
def get_formatting_funcs(typ):
    if typ == 'sheets':
        return [None, transDetails_formatting_sheets, shareProfitLoss_formatting_sheets, dailyProfitLoss_formatting_sheets, taxation_formatting_sheets]
    else:
        return [None, transDetails_formatting_excel, shareProfitLoss_formatting_excel, dailyProfitLoss_formatting_excel, taxation_formatting_excel]

def get_updating_func(typ):
    if typ == 'sheets':
        return update_sheet
    else:
        return update_excel

def update_sheet(spreadsheet, sheet_name, data, formatting_function=None):
    sheet = initialize_sheets(spreadsheet, sheet_name)
    display_and_format_sheets(sheet, data)
    if formatting_function is not None:
        formatting_function(spreadsheet, sheet)
    print(f"{sheet_name} updated Successfully!")

def update_excel(spreadsheet, sheet_name, data, formatting_function=None):
    sheet = initialize_excel(spreadsheet, sheet_name)
    display_and_format_excel(sheet, data)
    if formatting_function is not None:
        formatting_function(spreadsheet, sheet)
    print(f"{sheet_name} updated Successfully!")
    
def is_long_term(buy_date, sell_date):
    buy_date = datetime.strptime(buy_date, DATE_FORMAT)
    sell_date = datetime.strptime(sell_date, DATE_FORMAT)
    
    # Calculate the difference in days
    delta_days = (sell_date - buy_date).days
    
    # Check if the difference is 365 days or more
    return delta_days >= 365