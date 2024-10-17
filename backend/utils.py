from datetime import datetime
import gspread, requests, os
from constants import *
import pandas as pd
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
import openpyxl
import json
import numpy as np

# Functions for sheets
def authenticate_and_get_sheets(credentials_file, spreadsheet_id, credentials=None):
    print("Authenticating Sheets")
    if credentials is not None:
        try:
            credentials_obj = Credentials.from_authorized_user_info(credentials)
            if credentials_obj and credentials_obj.expired and credentials_obj.refresh_token:
                credentials_obj.refresh(Request())
            else:
                raise ValueError("OAuth Credentials are not available or invalid")
            
            gc = gspread.authorize(credentials_obj)
            spreadsheet = gc.open_by_key(spreadsheet_id)
            return spreadsheet
        except Exception as e:
            print(f"Unable to authorize spreadsheet {e}, exiting...")
            exit()
            
    else:
        gc = gspread.service_account(filename=credentials_file)
        spreadsheet = gc.open_by_key(spreadsheet_id)
        return spreadsheet
    
def read_data_from_sheets(spreadsheet, sheet_name):
    sheet = spreadsheet.worksheet(sheet_name)
    data = sheet.get_all_values()
    if len(data) == 0:
        return pd.DataFrame()
    df = pd.DataFrame(data[1:], columns=data[0])
    return df

def format_background_sheets(spreadsheet ,sheet, cell_range):
    requests = [{
        "updateCells": {
            "range": {
                "sheetId": sheet.id,
                "startRowIndex": int(cell_range.split(':')[0][1:]) - 1,
                "endRowIndex": int(cell_range.split(':')[1][1:]),
                "startColumnIndex": ord(cell_range.split(':')[0][0].upper()) - 65,
                "endColumnIndex": ord(cell_range.split(':')[1][0].upper()) - 64,
            },
            "fields": "userEnteredFormat.backgroundColor",
        }
    }]
    spreadsheet.batch_update({"requests":requests})

def initialize_sheets(spreadsheet, sheet_name):
    sheet = spreadsheet.worksheet(sheet_name)
    sheet.clear() 
    format_background_sheets(spreadsheet, sheet, CELL_RANGE)    
    return sheet

def update_sheet(spreadsheet, sheet_name, data, formatting_function=None):
    sheet = initialize_sheets(spreadsheet, sheet_name)
    display_and_format_sheets(sheet, data)
    if formatting_function is not None:
        formatting_function(spreadsheet, sheet)
    print(f"{sheet_name} updated Successfully!")

def replace_out_of_range_floats(obj):
    if isinstance(obj, float):
        if np.isnan(obj) or np.isinf(obj):
            return None
    elif isinstance(obj, list):
        return [replace_out_of_range_floats(item) for item in obj]
    elif isinstance(obj, dict):
        return {key: replace_out_of_range_floats(value) for key, value in obj.items()}
    return obj

def display_and_format_sheets(sheet, data):
    numeric_cols = data.select_dtypes(include=[np.number]).columns.tolist()
    data[numeric_cols] = data[numeric_cols].apply(pd.to_numeric, errors='coerce')  # Convert string to numeric
    data[numeric_cols] = data[numeric_cols].round(4)
    
    headers = data.columns.tolist() 
    data = data.values.tolist() 

    sheet.update('A1', [headers])
    sheet.insert_rows(data, 2)
    
    # Add formatting to headers
    num_columns = len(headers)
    header_range = f'A1:{chr(64 + num_columns)}1'
    sheet.format(header_range, {"textFormat": {"bold": True}})
    
    # Format numeric columns with 4 decimal places
    numeric_header_range = [f"{chr(65 + i)}1" for i, col in enumerate(headers) if col in numeric_cols]
    number_format = {
        "numberFormat": {
            "type": "NUMBER",
            "pattern": "#,##0"
        }
    }
    sheet.format(','.join(numeric_header_range), number_format)

def get_sheets_and_data(typ, credentials_file, spreadsheet_id, spreadsheet_file, credentials=None):
    if typ == 'sheets':
        spreadsheet = authenticate_and_get_sheets(
            credentials_file, spreadsheet_id, credentials)
        worksheets = spreadsheet.worksheets()
        sheet_names = [worksheet.title for worksheet in worksheets]
        raw_data = read_data_from_sheets(spreadsheet, sheet_names[0])
    else:
        spreadsheet = openpyxl.load_workbook(spreadsheet_file)
        sheet_names = spreadsheet.sheetnames
        raw_data = read_data_from_excel(spreadsheet, sheet_names[0])

    return spreadsheet, sheet_names, raw_data


# Functions for Excel
def get_symbol(row):
    symbol = row[Data_constants.NAME]
    symbol = str(symbol).split('-')[0]
    return str(symbol)

def read_data_from_excel(spreadsheet_file, sheet_name):
    df = pd.read_excel(spreadsheet_file, sheet_name=sheet_name)
    return df

def update_excel(spreadsheet, sheet_name, data, formatting_function=None):
    sheet = initialize_excel(spreadsheet, sheet_name)
    display_and_format_excel(sheet, data)
    if formatting_function is not None:
        formatting_function(spreadsheet, sheet)
    print(f"{sheet_name} updated Successfully!")
    
def display_and_format_excel(sheet, data):
    data = data.round(4)
    try:
        data[Raw_constants.DATE] = data[Raw_constants.DATE].dt.strftime(DATE_FORMAT)
    except Exception as e:
        pass
    headers = data.columns.tolist() 
    data = data.values.tolist() 
    sheet.append(headers)
    for row in data:
        sheet.append(row)
    for cell in sheet['1']:
        cell.font = openpyxl.styles.Font(bold=True)

def format_background_excel(sheet, cell_range):
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(fill_type=None)

def initialize_excel(spreadsheet, sheet_name):
    sheet = spreadsheet[sheet_name]
    sheet.clear()
    format_background_excel(sheet, CELL_RANGE)
    return sheet


# Common Functions
def get_updating_func(typ):
    if typ == 'sheets':
        return update_sheet
    else:
        return update_excel

def data_already_exists(raw_data, input_data):
    input_data.reset_index(drop=True, inplace=True)
    if not raw_data.empty:
        is_duplicate = raw_data[
            (raw_data[Raw_constants.DATE] == input_data[Raw_constants.DATE][0]) &
            (raw_data[Raw_constants.NAME] == input_data[Raw_constants.NAME][0])
        ].shape[0] > 0

        if is_duplicate:
            print("Orders Data Already Exists in the file, Exiting...")
            exit()

def get_data_date(date):
    date_obj = datetime.strptime(date, DATA_TIME_FORMAT)
    date_str = datetime.strftime(date_obj, DATE_FORMAT)
    return date_str

def get_data_quantity(row):
    quantity = row[Data_constants.QUANTITY]
    val = str(quantity).split('.')[0]
    if str(row[Data_constants.TYPE]).upper() == SELL:
        val = '-' + val
    return val

def get_net_amount(row):
    val = int(row[Raw_constants.QUANTITY]) * float(row[Raw_constants.PRICE])
    return str(val)

def format_add_data(input_data):
    input_data[Data_constants.QUANTITY] = input_data.apply(
        get_data_quantity, axis=1)
    
    constants_dict = {key: value for key, value in vars(Raw_constants).items(
    ) if (isinstance(value, str) and not value.startswith('python'))}
    df = pd.DataFrame(columns=list(constants_dict.values()))
    
    df[Raw_constants.DATE] = input_data[Data_constants.DATE].apply(
        lambda x: get_data_date(x))
    df[Raw_constants.NAME] = input_data.apply(get_symbol, axis=1)
    df[Raw_constants.PRICE] = input_data[Data_constants.PRICE]
    df[Raw_constants.QUANTITY] = input_data[Data_constants.QUANTITY]
    df[Raw_constants.NET_AMOUNT] = df.apply(get_net_amount, axis=1)
    df[Raw_constants.STOCK_EXCHANGE] = input_data[Data_constants.STOCK_EXCHANGE]
    return df

def check_valid_path(path):
    if path is None or not os.path.exists(path):
        return None
    return True

def get_valid_path(path):
    if path is None:
        return None
    if not os.path.exists(path):
        print("The Provided Path for file downloaded from Zerodha does not exist!")
        path = input("Enter correct Absolute Path, including /home: ")
        path = get_valid_path(path)   
    return path

def update_env_file(key, value, env_file):
    print("updating env file")
    with open(env_file, 'r') as file:
        lines = file.readlines()    

    updated_lines = []
    for line in lines:
        if line.strip().startswith(f"{key}="):
            line = f"{key}={value}\n"
        updated_lines.append(line)

    print("Lines updated are: ", updated_lines)
    with open(env_file, 'w') as file:
        file.writelines(updated_lines)

def get_args_and_input(args, excel_file_name, spreadsheet_id, env_file):
    
    input_file = None
    value = ''
    key = 'EXCEL_FILE_NAME'
    credentials=None
    print("Args length: ", len(args), "and Args: ", args)
    if len(args) > 4:
        credentials = json.loads(args[4])       
    if len(args) > 3:
        value = args[3]
    if len(args) > 2:
        typ = args[2].lower()
        if typ == 'sheets':
            key = 'SPREADSHEET_ID'
        input_file = args[1]
        if input_file == 'None':
            input_file = None
    else:
        if len(args) > 1:
            input_file = args[1]
        else:
            input_file = input("Please enter the absolute path of the file downloaded from Zerodha: ")
        print("\nPlease select 'excel' or 'sheets, leave empty to use excel as default")
        typ = input("Enter your choice: ")
        typ = typ.lower()
        if (typ == 'excel' or typ == ''):
            print(
                f"{excel_file_name} is the default file, enter name below if you wish to change it, leave empty otherwise")
            value = input("Enter your choice: ")

        elif (typ== 'sheets'):
            print(f"{spreadsheet_id} is the default google sheet, enter spreadsheet_id below if you wish to change it, leave empty otherwise")
            value = input("Enter your choice: ")
            key = 'SPREADSHEET_ID'

    if value is not None and value != '':
        update_env_file(key, value, env_file)

    return input_file, typ, credentials

def credentials_to_dict(credentials):
    return {
        'token': credentials.token,
        'refresh_token': credentials.refresh_token,
        'token_uri': credentials.token_uri,
        'client_id': credentials.client_id,
        'client_secret': credentials.client_secret,
        'scopes': credentials.scopes
    }