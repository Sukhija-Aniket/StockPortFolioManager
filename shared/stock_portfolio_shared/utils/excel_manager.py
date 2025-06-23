"""
Excel utilities for Stock Portfolio Manager
"""

import openpyxl
import pandas as pd
import logging
from openpyxl.styles import Font, PatternFill
from stock_portfolio_shared.models.spreadsheet_task import SpreadsheetTask
from stock_portfolio_shared.utils.base_manager import BaseManager
from ..constants.general_constants import BUY, CELL_RANGE, DATA_TIME_FORMAT, SELL
from ..constants.raw_constants import Raw_constants
from ..utils.data_processor import DataProcessor
import os

logger = logging.getLogger(__name__)

class ExcelManager(BaseManager):
    """Manages Excel operations"""
    
    def __init__(self):
        self.data_processor = DataProcessor()
    
    def load_workbook(self, spreadsheet_file):
        """Load Excel workbook"""
        return openpyxl.load_workbook(spreadsheet_file)
    
    def get_sheet_names(self, spreadsheet_task: SpreadsheetTask):
        """Get sheet names from Excel file"""
        spreadsheet = self.get_spreadsheet(spreadsheet_task)
        return spreadsheet.sheetnames
    
    def read_data(self, spreadsheet, sheet_name):
        """Read data from Excel file using spreadsheet object"""
        sheet_names = spreadsheet.sheetnames
        if sheet_name not in sheet_names:
            raise ValueError(f"Sheet {sheet_name} not found in spreadsheet")
        
        # Get the worksheet from the spreadsheet object
        worksheet = spreadsheet[sheet_name]
        
        # Convert worksheet data to DataFrame
        data = []
        headers = []
        
        # Get headers from first row
        for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = row
            break
        
        # Get data rows (skip header row)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):  # Skip empty rows
                data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        return df
    
    def format_background_excel(self, sheet, cell_range):
        """Format background of Excel sheet"""
        for row in sheet[cell_range]:
            for cell in row:
                cell.fill = PatternFill(fill_type=None)
    
    def display_and_format_excel(self, sheet, data):
        """Display and format data in Excel"""
        data = data.round(4)
        try:
            data[Raw_constants.DATE] = data[Raw_constants.DATE].dt.strftime(DATA_TIME_FORMAT)
        except Exception as e:
            pass
        headers = data.columns.tolist() 
        data = data.values.tolist() 
        sheet.append(headers)
        for row in data:
            sheet.append(row)
        for cell in sheet['1']:
            cell.font = Font(bold=True)     
    
    def add_data(self, input_data, spreadsheet, sheet_name, allow_duplicates=False, formatting_function=None):
        """Upload data to Excel file"""
        try:
            raw_data = self.read_data(spreadsheet, sheet_name)
            validated_input_data = self.validate_data(raw_data, input_data)
            
            if not allow_duplicates:
                if self.data_processor.data_already_exists(raw_data, validated_input_data):
                    logger.info("Data already exists in Excel")
                    return
            raw_data = pd.concat([raw_data, validated_input_data], ignore_index=True)
            self.update_data(spreadsheet, sheet_name, raw_data, formatting_function)
            logger.info("Data uploaded successfully to Excel")
        except Exception as e:
            logger.error(f"Error uploading data to Excel: {e}")
            raise
            
    
    def get_spreadsheet(self, spreadsheet_task: SpreadsheetTask):
        spreadsheet = self.load_workbook(spreadsheet_task.spreadsheet_id)
        return spreadsheet

    
    def _initialize_excel(self, spreadsheet, sheet_name):
        """Initialize Excel sheet"""
        sheet = spreadsheet[sheet_name]
        sheet.clear()
        self.format_background_excel(sheet, CELL_RANGE)
        return sheet
    
    def update_data(self, spreadsheet, sheet_name, data, formatting_function=None):
        """Update Excel with data"""
        sheet = self._initialize_excel(spreadsheet, sheet_name)
        self.display_and_format_excel(sheet, data)
        if formatting_function is not None:
            formatting_function(spreadsheet, sheet)
        logger.info(f"{sheet_name} updated Successfully!")
    
    
    def transDetails_formatting(self, sheet):
        cell_range = CELL_RANGE
        redFill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        blueFill = openpyxl.styles.PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
        for row in sheet[cell_range]:
            if row[0].value is None or row[0].value == '':
                break
            if row[4].value == BUY:
                for cell in row:
                    cell.fill = blueFill
            elif row[4].value == SELL:
                for cell in row:
                    cell.fill = redFill

    def shareProfitLoss_formatting(self, sheet):
        cell_range = CELL_RANGE
        redFill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        blueFill = openpyxl.styles.PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
        for row in sheet[cell_range]:
            if row[0].value is None or row[0].value == '':
                break
            remaining_shares = pd.to_numeric(row[7].value, errors='coerce', thousands=',')
            profit = pd.to_numeric(row[9].value, errors='coerce', thousands=',')
            if remaining_shares == 0:
                if profit >= 0.0:
                    for cell in row:
                        cell.fill = blueFill
                else:
                    for cell in row:
                        cell.fill = redFill

    def dailyProfitLoss_formatting(self, sheet):
        cell_range = CELL_RANGE
        redFill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        blueFill = openpyxl.styles.PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
        for row in sheet[cell_range]:
            if row[0].value is None or row[0].value == '':
                break
            if row[9].value != '':
                if pd.to_numeric(row[9].value, errors='coerce', thousands=',') > 0.0:
                    for cell in row:
                        cell.fill = blueFill
                else :
                    for cell in row:
                        cell.fill = redFill
                    
    def taxation_formatting(self, sheet):
        cell_range = CELL_RANGE
        redFill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        blueFill = openpyxl.styles.PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
        for row in sheet[cell_range]:
            if row[0].value is None or row[0].value == '':
                break
            if row[9].value != '':
                if pd.to_numeric(row[2].value, errors='coerce', thousands=',') + pd.to_numeric(row[3].value, errors='coerce', thousands=',') + pd.to_numeric(row[4].value, errors='coerce', thousands=',') > 0.0:
                    for cell in row:
                        cell.fill = blueFill
                elif pd.to_numeric(row[2].value, errors='coerce', thousands=',') + pd.to_numeric(row[3].value, errors='coerce', thousands=',') + pd.to_numeric(row[4].value, errors='coerce', thousands=',') < 0.0:
                    for cell in row:
                        cell.fill = redFill
                        
    def get_formatting_funcs(self, sheet_names):
        """Get formatting functions"""
        return {
            sheet_names[0]: None,
            sheet_names[1]: self.transDetails_formatting,
            sheet_names[2]: self.shareProfitLoss_formatting,
            sheet_names[3]: self.dailyProfitLoss_formatting,
            sheet_names[4]: self.taxation_formatting
        }