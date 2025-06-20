"""
Excel utilities for Stock Portfolio Manager
"""

import openpyxl
import pandas as pd
import logging
from openpyxl.styles import Font, PatternFill
from ..constants import CELL_RANGE, Raw_constants, DATE_FORMAT

logger = logging.getLogger(__name__)

class ExcelManager:
    """Manages Excel operations"""
    
    def load_workbook(self, spreadsheet_file):
        """Load Excel workbook"""
        return openpyxl.load_workbook(spreadsheet_file)
    
    def read_data_from_excel(self, spreadsheet_file, sheet_name):
        """Read data from Excel file"""
        df = pd.read_excel(spreadsheet_file, sheet_name=sheet_name)
        return df
    
    def format_background_excel(self, sheet, cell_range):
        """Format background of Excel sheet"""
        for row in sheet[cell_range]:
            for cell in row:
                cell.fill = PatternFill(fill_type=None)
    
    def display_and_format_excel(self, sheet, data):
        """Display and format data in Excel"""
        data = data.round(4)
        try:
            data[Raw_constants.DATE] = data[Raw_constants.DATE].dt.strftime(DATE_FORMAT)
        except Exception as e:
            pass
        headers = data.columns.tolist() 
        data = data.values.tolist() 
        sheet.append(headers)
        for row in data:
            sheet.append(row)
        for cell in sheet['1']:
            cell.font = Font(bold=True)
    
    def initialize_excel(self, spreadsheet, sheet_name):
        """Initialize Excel sheet"""
        sheet = spreadsheet[sheet_name]
        sheet.clear()
        self.format_background_excel(sheet, CELL_RANGE)
        return sheet
    
    def update_excel(self, spreadsheet, sheet_name, data, formatting_function=None):
        """Update Excel with data"""
        sheet = self.initialize_excel(spreadsheet, sheet_name)
        self.display_and_format_excel(sheet, data)
        if formatting_function is not None:
            formatting_function(spreadsheet, sheet)
        logger.info(f"{sheet_name} updated Successfully!")
    
    def get_updating_func(self, typ):
        """Get updating function based on type"""
        if typ == 'sheets':
            from .sheets import SheetsManager
            sheets_manager = SheetsManager()
            return sheets_manager.update_sheet
        else:
            return self.update_excel 